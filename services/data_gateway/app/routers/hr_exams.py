"""HR MCQ exam authoring + results — HR workflow Phase 2.

An HR manager authors an exam (MCQ questions + pass threshold), publishes it,
assigns it to applicants (minting an opaque magic-link the applicant uses to take
it WITHOUT logging in), and reviews graded results.

MULTI-TENANT: every endpoint is scoped to the caller's company_id (reusing the
Phase-1 get_hr_company/HrCtxDep). An HR can NEVER see or touch another company's
exams — all reads/writes filter by company_id, so a cross-company id returns 404.

SECURITY:
  - correct_index is returned ONLY to HR (authoring/detail/breakdown) — never on
    the applicant take path (that lives in exam_take.py).
  - Questions LOCK (409) once any attempt exists, so a graded exam can't be
    silently changed under candidates.
  - Magic-link mint returns the RAW token exactly once; only its HMAC hash is
    stored. Re-assigning rotates the link (revokes the prior active one).
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, datetime, timedelta
from typing import Annotated, Any

import structlog
from fastapi import APIRouter, File, HTTPException, Query, Response, UploadFile, status
from pydantic import BaseModel, Field, ValidationError, field_validator, model_validator
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.exam_ai_client import ExamGenerationError, generate_exam_questions_remote
from app.exam_grading import GradeInput, GradeQuestion, grade_breakdown
from app.exam_link import hash_exam_token, mint_exam_token
from app.mailer import enqueue_email
from app.models import (
    Applicant,
    CodingQuestion,
    Exam,
    ExamAssignment,
    ExamAttempt,
    ExamQuestion,
    ExamRound,
    ExamSection,
)
from app.routers.hr_applicants import DbSessionDep, HrCtxDep

log = structlog.get_logger(__name__)

router = APIRouter(prefix="/hr", tags=["hr-exams"])

_VALID_STATUSES = {"draft", "published", "closed"}


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------
class ExamCreateIn(BaseModel):
    title: str = Field(min_length=1, max_length=300)
    description: str | None = None
    target_job_title: str | None = None
    pass_threshold: int = Field(default=60, ge=0, le=100)
    time_limit_seconds: int | None = Field(default=None, ge=10, le=86_400)
    allow_retake: bool = False
    # 'mcq' (default) or 'coding' — selects the question type + grader. This also
    # becomes the kind of the auto-created default section (back-compat).
    kind: str = Field(default="mcq")
    # When True, passing the terminal round auto-creates a scheduled interview
    # invite + emails the candidate; when False HR invites manually.
    auto_advance_on_pass: bool = False

    @field_validator("kind")
    @classmethod
    def _validate_kind(cls, v: str) -> str:
        v = (v or "mcq").lower()
        if v not in {"mcq", "coding"}:
            raise ValueError("kind must be mcq | coding")
        return v


class ExamUpdateIn(BaseModel):
    title: str | None = Field(default=None, min_length=1, max_length=300)
    description: str | None = None
    target_job_title: str | None = None
    pass_threshold: int | None = Field(default=None, ge=0, le=100)
    time_limit_seconds: int | None = Field(default=None, ge=10, le=86_400)
    allow_retake: bool | None = None
    auto_advance_on_pass: bool | None = None
    status: str | None = None

    @field_validator("status")
    @classmethod
    def _validate_status(cls, v: str | None) -> str | None:
        if v is not None and v not in _VALID_STATUSES:
            raise ValueError(f"status must be one of {sorted(_VALID_STATUSES)}")
        return v


class QuestionIn(BaseModel):
    prompt: str = Field(min_length=1, max_length=2000)
    options: list[str] = Field(min_length=2, max_length=6)
    correct_index: int = Field(ge=0)
    points: int = Field(default=1, ge=1, le=100)

    @field_validator("options")
    @classmethod
    def _validate_options(cls, v: list[str]) -> list[str]:
        cleaned = [o.strip() for o in v]
        if any(not o for o in cleaned):
            raise ValueError("options must be non-empty strings")
        return cleaned

    @model_validator(mode="after")
    def _validate_correct_index(self) -> QuestionIn:
        if self.correct_index >= len(self.options):
            raise ValueError("correct_index out of range for options")
        return self


class QuestionUpdateIn(BaseModel):
    prompt: str | None = Field(default=None, min_length=1, max_length=2000)
    options: list[str] | None = Field(default=None, min_length=2, max_length=6)
    correct_index: int | None = Field(default=None, ge=0)
    points: int | None = Field(default=None, ge=1, le=100)

    @field_validator("options")
    @classmethod
    def _validate_options(cls, v: list[str] | None) -> list[str] | None:
        if v is None:
            return v
        cleaned = [o.strip() for o in v]
        if any(not o for o in cleaned):
            raise ValueError("options must be non-empty strings")
        return cleaned


class ReorderIn(BaseModel):
    question_ids: list[uuid.UUID] = Field(min_length=1)


class BulkQuestionsIn(BaseModel):
    """Insert many questions at once (used by AI-generate 'add all' + Excel import)."""

    questions: list[QuestionIn] = Field(min_length=1, max_length=200)


class GenerateQuestionsIn(BaseModel):
    """Ask Gemini (via feedback_billing) to draft MCQs — returned for preview, NOT saved."""

    topic: str = Field(min_length=1, max_length=300)
    num_questions: int = Field(default=5, ge=1, le=30)
    difficulty: str = Field(default="medium")
    language: str = Field(default="en")

    @field_validator("difficulty")
    @classmethod
    def _validate_difficulty(cls, v: str) -> str:
        v = (v or "medium").lower()
        if v not in {"easy", "medium", "hard", "mixed"}:
            raise ValueError("difficulty must be easy | medium | hard | mixed")
        return v

    @field_validator("language")
    @classmethod
    def _validate_language(cls, v: str) -> str:
        v = (v or "en").lower()
        if v not in {"en", "hi", "te"}:
            raise ValueError("language must be en | hi | te")
        return v


class GeneratedQuestionOut(BaseModel):
    """A drafted question — no id/position (it isn't persisted until HR adds it)."""

    prompt: str
    options: list[str]
    correct_index: int
    points: int = 1


class GenerateQuestionsOut(BaseModel):
    questions: list[GeneratedQuestionOut]


class ImportRowError(BaseModel):
    row: int
    message: str


class ImportQuestionsOut(BaseModel):
    added: int
    errors: list[ImportRowError]
    questions: list[QuestionOut]


class AssignIn(BaseModel):
    applicant_ids: list[uuid.UUID] = Field(min_length=1, max_length=200)
    ttl_hours: int | None = Field(default=None, ge=1, le=8760)
    # Optional: assign a SPECIFIC round (defaults to the exam's first round when
    # omitted — the back-compat single-round path). Optional scheduled start time.
    round_id: uuid.UUID | None = None
    scheduled_at: datetime | None = None


class QuestionOut(BaseModel):
    """HR-facing question — INCLUDES correct_index (HR only; never on take path)."""

    id: str
    prompt: str
    options: list[str]
    correct_index: int
    points: int
    position: int


class ExamOut(BaseModel):
    id: str
    title: str
    description: str | None
    target_job_title: str | None
    pass_threshold: int
    time_limit_seconds: int | None
    allow_retake: bool
    status: str
    kind: str
    auto_advance_on_pass: bool
    created_at: str


class ExamSummaryOut(ExamOut):
    question_count: int
    attempt_count: int


class ExamDetailOut(ExamOut):
    questions: list[QuestionOut]
    attempt_count: int


class AssignOut(BaseModel):
    assignment_id: str
    applicant_id: str
    applicant_name: str
    magic_link: str  # raw token embedded — returned ONCE, at mint time only
    expires_at: str
    status: str
    round_id: str | None = None
    scheduled_at: str | None = None


class AssignmentOut(BaseModel):
    assignment_id: str
    applicant_id: str
    applicant_name: str
    status: str
    expires_at: str
    consumed_at: str | None
    created_at: str


class AttemptResultOut(BaseModel):
    attempt_id: str
    applicant_id: str
    applicant_name: str
    score_raw: int | None
    score_max: int | None
    score_percent: int | None
    passed: bool | None
    status: str
    submitted_at: str | None
    attempt_no: int


# ---------------------------------------------------------------------------
# Helpers (tenant isolation lives here)
# ---------------------------------------------------------------------------
async def _get_owned_exam(db: AsyncSession, company_id: uuid.UUID, exam_id: uuid.UUID) -> Exam:
    exam = await db.scalar(
        select(Exam).where(
            Exam.id == exam_id, Exam.company_id == company_id, Exam.deleted_at.is_(None)
        )
    )
    if exam is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Exam not found.")
    return exam


async def _create_default_round_section(
    db: AsyncSession, exam: Exam, *, kind: str
) -> tuple[ExamRound, ExamSection]:
    """Create the exam's default Round 1 + one section (mirrors the migration
    backfill), so legacy exam-scoped question/assignment flows keep working on a
    fresh exam. Caller owns the commit. advances_to_interview=True — the single
    round is terminal until HR adds more."""
    now = datetime.now(tz=UTC)
    rnd = ExamRound(
        id=uuid.uuid4(),
        exam_id=exam.id,
        company_id=exam.company_id,
        round_number=1,
        title="Round 1",
        pass_threshold=exam.pass_threshold,
        time_limit_seconds=exam.time_limit_seconds,
        advances_to_interview=True,
        status="draft",
        position=1,
        created_at=now,
        updated_at=now,
    )
    db.add(rnd)
    sec = ExamSection(
        id=uuid.uuid4(),
        round_id=rnd.id,
        exam_id=exam.id,
        company_id=exam.company_id,
        title="Section 1",
        kind=kind,
        time_limit_seconds=None,
        position=1,
        created_at=now,
        updated_at=now,
    )
    db.add(sec)
    return rnd, sec


async def _default_round(
    db: AsyncSession, company_id: uuid.UUID, exam_id: uuid.UUID
) -> ExamRound:
    """The exam's first live round (back-compat single-round path)."""
    rnd = await db.scalar(
        select(ExamRound)
        .where(
            ExamRound.exam_id == exam_id,
            ExamRound.company_id == company_id,
            ExamRound.deleted_at.is_(None),
        )
        .order_by(ExamRound.position.asc())
        .limit(1)
    )
    if rnd is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Exam has no rounds.")
    return rnd


async def _default_section(
    db: AsyncSession, company_id: uuid.UUID, exam_id: uuid.UUID, kind: str
) -> ExamSection:
    """The exam's first live section of the given kind (back-compat path: legacy
    exams have exactly one section, of kind == exam.kind)."""
    sec = await db.scalar(
        select(ExamSection)
        .where(
            ExamSection.exam_id == exam_id,
            ExamSection.company_id == company_id,
            ExamSection.kind == kind,
            ExamSection.deleted_at.is_(None),
        )
        .order_by(ExamSection.position.asc())
        .limit(1)
    )
    if sec is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Exam has no {kind} section.",
        )
    return sec


async def _question_count(db: AsyncSession, exam: Exam) -> int:
    """Count live questions in the table that matches the exam's kind."""
    model = CodingQuestion if exam.kind == "coding" else ExamQuestion
    n = await db.scalar(
        select(func.count()).select_from(model).where(
            model.exam_id == exam.id, model.deleted_at.is_(None)
        )
    )
    return int(n or 0)


async def _attempt_count(db: AsyncSession, company_id: uuid.UUID, exam_id: uuid.UUID) -> int:
    n = await db.scalar(
        select(func.count()).select_from(ExamAttempt).where(
            ExamAttempt.exam_id == exam_id,
            ExamAttempt.company_id == company_id,
            ExamAttempt.deleted_at.is_(None),
        )
    )
    return int(n or 0)


async def _live_questions(
    db: AsyncSession, company_id: uuid.UUID, exam_id: uuid.UUID
) -> list[ExamQuestion]:
    rows = (
        await db.execute(
            select(ExamQuestion)
            .where(
                ExamQuestion.exam_id == exam_id,
                ExamQuestion.company_id == company_id,
                ExamQuestion.deleted_at.is_(None),
            )
            .order_by(ExamQuestion.position.asc())
        )
    ).scalars().all()
    return list(rows)


async def _get_owned_question(
    db: AsyncSession, company_id: uuid.UUID, exam_id: uuid.UUID, qid: uuid.UUID
) -> ExamQuestion:
    q = await db.scalar(
        select(ExamQuestion).where(
            ExamQuestion.id == qid,
            ExamQuestion.exam_id == exam_id,
            ExamQuestion.company_id == company_id,
            ExamQuestion.deleted_at.is_(None),
        )
    )
    if q is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found.")
    return q


async def _require_no_attempts(
    db: AsyncSession, company_id: uuid.UUID, exam_id: uuid.UUID
) -> None:
    """Questions are immutable once any attempt exists (graded-exam integrity)."""
    if await _attempt_count(db, company_id, exam_id) > 0:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="This exam already has attempts — its questions are locked.",
        )


def _exam_out(e: Exam) -> ExamOut:
    return ExamOut(
        id=str(e.id),
        title=e.title,
        description=e.description,
        target_job_title=e.target_job_title,
        pass_threshold=e.pass_threshold,
        time_limit_seconds=e.time_limit_seconds,
        allow_retake=e.allow_retake,
        status=e.status,
        kind=e.kind,
        auto_advance_on_pass=e.auto_advance_on_pass,
        created_at=e.created_at.isoformat(),
    )


def _question_out(q: ExamQuestion) -> QuestionOut:
    return QuestionOut(
        id=str(q.id),
        prompt=q.prompt,
        options=list(q.options or []),
        correct_index=q.correct_index,
        points=q.points,
        position=q.position,
    )


async def _bulk_insert_questions(
    db: AsyncSession,
    company_id: uuid.UUID,
    section: ExamSection,
    items: list[QuestionIn],
) -> list[ExamQuestion]:
    """Append many already-validated questions at sequential positions WITHIN a
    section. Caller must have checked ownership + the attempt-lock first."""
    max_pos = await db.scalar(
        select(func.max(ExamQuestion.position)).where(
            ExamQuestion.section_id == section.id, ExamQuestion.deleted_at.is_(None)
        )
    )
    next_pos = (int(max_pos) + 1) if max_pos is not None else 0
    now = datetime.now(tz=UTC)
    created: list[ExamQuestion] = []
    for offset, item in enumerate(items):
        q = ExamQuestion(
            id=uuid.uuid4(),
            exam_id=section.exam_id,
            section_id=section.id,
            company_id=company_id,
            prompt=item.prompt.strip(),
            options=item.options,
            correct_index=item.correct_index,
            points=item.points,
            position=next_pos + offset,
            created_at=now,
            updated_at=now,
        )
        db.add(q)
        created.append(q)
    await db.commit()
    return created


# --- Spreadsheet (Excel / CSV) bulk-import helpers --------------------------
# Template layout (one row per question, header in row 1):
#   Question | Option A | Option B | Option C | Option D | Correct | Points
# "Correct" names the right option by letter (A-D), number (1-4), or its text.
_TEMPLATE_HEADER: list[str] = [
    "Question", "Option A", "Option B", "Option C", "Option D", "Correct", "Points"
]
_MAX_IMPORT_BYTES: int = 2_000_000  # 2 MB upload cap (DoS guard)


def _correct_to_index(raw: str, options: list[str]) -> int:
    """Resolve the 'Correct' cell to a 0-based option index. Raises ValueError."""
    s = (raw or "").strip()
    if not s:
        raise ValueError("missing correct answer")
    if len(s) == 1 and s.isalpha():  # 'A'..'D'
        idx = ord(s.upper()) - ord("A")
        if 0 <= idx < len(options):
            return idx
        raise ValueError(f"correct '{s}' is out of range for {len(options)} options")
    if s.isdigit():  # '1'..'4'
        idx = int(s) - 1
        if 0 <= idx < len(options):
            return idx
        raise ValueError(f"correct '{s}' is out of range for {len(options)} options")
    for i, o in enumerate(options):  # literal option text
        if o.strip().casefold() == s.casefold():
            return i
    raise ValueError(f"correct answer '{s}' matches no option")


def _read_spreadsheet(filename: str, content: bytes) -> list[list[str]]:
    """Read an uploaded .xlsx or .csv into a list of string rows."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return [[(c or "") for c in row] for row in csv.reader(io.StringIO(text))]
    try:
        import openpyxl  # lazy: only needed for the Excel path
    except ImportError as exc:  # pragma: no cover - dep is in requirements
        raise HTTPException(
            status_code=500, detail="Excel support is not installed on the server."
        ) from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises many types on bad files
        raise HTTPException(
            status_code=400, detail="Could not read that file — is it a valid .xlsx or .csv?"
        ) from exc
    ws = wb.active
    rows: list[list[str]] = []
    if ws is not None:
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
    wb.close()
    return rows


def _parse_question_rows(
    rows: list[list[str]],
) -> tuple[list[QuestionIn], list[ImportRowError]]:
    """Parse spreadsheet rows into QuestionIn objects + a per-row error list."""
    items: list[QuestionIn] = []
    errors: list[ImportRowError] = []
    start = 0
    if rows and rows[0] and "question" in (rows[0][0] or "").strip().casefold():
        start = 1  # skip the header row
    for i in range(start, len(rows)):
        rownum = i + 1  # 1-based for human-facing messages
        cells = [(c or "").strip() for c in rows[i]]
        if not any(cells):
            continue  # blank line — skip silently
        prompt = cells[0] if cells else ""
        options = [cells[j] for j in range(1, 5) if j < len(cells) and cells[j]]
        correct_raw = cells[5] if len(cells) > 5 else ""
        points_raw = cells[6] if len(cells) > 6 else ""
        if not prompt:
            errors.append(ImportRowError(row=rownum, message="missing question text"))
            continue
        if len(options) < 2:
            errors.append(ImportRowError(row=rownum, message="need at least 2 options"))
            continue
        try:
            correct_index = _correct_to_index(correct_raw, options)
        except ValueError as exc:
            errors.append(ImportRowError(row=rownum, message=str(exc)))
            continue
        points = 1
        if points_raw:
            try:
                points = max(1, min(100, int(float(points_raw))))
            except ValueError:
                points = 1
        try:
            items.append(
                QuestionIn(
                    prompt=prompt, options=options, correct_index=correct_index, points=points
                )
            )
        except ValidationError as exc:
            errs = exc.errors()
            msg = str(errs[0].get("msg", "invalid question")) if errs else "invalid question"
            errors.append(ImportRowError(row=rownum, message=msg))
    return items, errors


# ---------------------------------------------------------------------------
# Exam CRUD
# ---------------------------------------------------------------------------
@router.post("/exams", status_code=status.HTTP_201_CREATED, response_model=ExamOut)
async def create_exam(body: ExamCreateIn, ctx: HrCtxDep, db: DbSessionDep) -> ExamOut:
    hr_uid, company_id = ctx
    now = datetime.now(tz=UTC)
    exam = Exam(
        id=uuid.uuid4(),
        company_id=company_id,
        created_by_user_id=hr_uid,
        title=body.title.strip(),
        description=body.description,
        target_job_title=body.target_job_title,
        pass_threshold=body.pass_threshold,
        time_limit_seconds=body.time_limit_seconds,
        allow_retake=body.allow_retake,
        status="draft",
        kind=body.kind,
        auto_advance_on_pass=body.auto_advance_on_pass,
        created_at=now,
        updated_at=now,
    )
    db.add(exam)
    await db.flush()  # exam.id available for the default round/section FKs
    await _create_default_round_section(db, exam, kind=body.kind)
    await db.commit()
    log.info("hr.exam.created", exam_id=str(exam.id), company_id=str(company_id), kind=body.kind)
    return _exam_out(exam)


@router.get("/exams", response_model=list[ExamSummaryOut])
async def list_exams(
    ctx: HrCtxDep,
    db: DbSessionDep,
    status_filter: Annotated[str | None, Query(alias="status")] = None,
) -> list[ExamSummaryOut]:
    _hr_uid, company_id = ctx
    stmt = select(Exam).where(Exam.company_id == company_id, Exam.deleted_at.is_(None))
    if status_filter:
        stmt = stmt.where(Exam.status == status_filter)
    stmt = stmt.order_by(Exam.created_at.desc())
    exams = (await db.execute(stmt)).scalars().all()

    out: list[ExamSummaryOut] = []
    for e in exams:
        qn = await _question_count(db, e)
        an = await _attempt_count(db, company_id, e.id)
        out.append(
            ExamSummaryOut(**_exam_out(e).model_dump(), question_count=qn, attempt_count=an)
        )
    return out


@router.get("/exams/{exam_id}", response_model=ExamDetailOut)
async def get_exam(exam_id: uuid.UUID, ctx: HrCtxDep, db: DbSessionDep) -> ExamDetailOut:
    _hr_uid, company_id = ctx
    exam = await _get_owned_exam(db, company_id, exam_id)
    questions = await _live_questions(db, company_id, exam_id)
    an = await _attempt_count(db, company_id, exam_id)
    return ExamDetailOut(
        **_exam_out(exam).model_dump(),
        questions=[_question_out(q) for q in questions],
        attempt_count=an,
    )


@router.patch("/exams/{exam_id}", response_model=ExamOut)
async def update_exam(
    exam_id: uuid.UUID, body: ExamUpdateIn, ctx: HrCtxDep, db: DbSessionDep
) -> ExamOut:
    _hr_uid, company_id = ctx
    exam = await _get_owned_exam(db, company_id, exam_id)

    if body.status == "published" and await _question_count(db, exam) < 1:
        raise HTTPException(
            status_code=400, detail="Add at least one question before publishing."
        )

    if body.title is not None:
        exam.title = body.title.strip()
    if body.description is not None:
        exam.description = body.description
    if body.target_job_title is not None:
        exam.target_job_title = body.target_job_title
    if body.pass_threshold is not None:
        exam.pass_threshold = body.pass_threshold
    if body.time_limit_seconds is not None:
        exam.time_limit_seconds = body.time_limit_seconds
    if body.allow_retake is not None:
        exam.allow_retake = body.allow_retake
    if body.auto_advance_on_pass is not None:
        exam.auto_advance_on_pass = body.auto_advance_on_pass
    if body.status is not None:
        exam.status = body.status
    exam.updated_at = datetime.now(tz=UTC)
    await db.commit()
    return _exam_out(exam)


# ---------------------------------------------------------------------------
# Questions (locked once attempts exist)
# ---------------------------------------------------------------------------
@router.post(
    "/exams/{exam_id}/questions", status_code=status.HTTP_201_CREATED, response_model=QuestionOut
)
async def add_question(
    exam_id: uuid.UUID, body: QuestionIn, ctx: HrCtxDep, db: DbSessionDep
) -> QuestionOut:
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    await _require_no_attempts(db, company_id, exam_id)
    # Back-compat: target the exam's default MCQ section.
    section = await _default_section(db, company_id, exam_id, "mcq")

    max_pos = await db.scalar(
        select(func.max(ExamQuestion.position)).where(
            ExamQuestion.section_id == section.id, ExamQuestion.deleted_at.is_(None)
        )
    )
    now = datetime.now(tz=UTC)
    q = ExamQuestion(
        id=uuid.uuid4(),
        exam_id=exam_id,
        section_id=section.id,
        company_id=company_id,
        prompt=body.prompt.strip(),
        options=body.options,
        correct_index=body.correct_index,
        points=body.points,
        position=(int(max_pos) + 1) if max_pos is not None else 0,
        created_at=now,
        updated_at=now,
    )
    db.add(q)
    await db.commit()
    return _question_out(q)


@router.patch("/exams/{exam_id}/questions/{qid}", response_model=QuestionOut)
async def update_question(
    exam_id: uuid.UUID, qid: uuid.UUID, body: QuestionUpdateIn, ctx: HrCtxDep, db: DbSessionDep
) -> QuestionOut:
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    await _require_no_attempts(db, company_id, exam_id)
    q = await _get_owned_question(db, company_id, exam_id, qid)

    new_options = body.options if body.options is not None else list(q.options or [])
    new_correct = body.correct_index if body.correct_index is not None else q.correct_index
    if new_correct >= len(new_options):
        raise HTTPException(status_code=400, detail="correct_index out of range for options.")

    if body.prompt is not None:
        q.prompt = body.prompt.strip()
    q.options = new_options
    q.correct_index = new_correct
    if body.points is not None:
        q.points = body.points
    q.updated_at = datetime.now(tz=UTC)
    await db.commit()
    return _question_out(q)


@router.delete("/exams/{exam_id}/questions/{qid}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_question(
    exam_id: uuid.UUID, qid: uuid.UUID, ctx: HrCtxDep, db: DbSessionDep
) -> Response:
    _hr_uid, company_id = ctx
    exam = await _get_owned_exam(db, company_id, exam_id)
    await _require_no_attempts(db, company_id, exam_id)
    q = await _get_owned_question(db, company_id, exam_id, qid)

    live = await _live_questions(db, company_id, exam_id)
    if exam.status == "published" and len(live) <= 1:
        raise HTTPException(
            status_code=400,
            detail="A published exam must keep at least one question. Unpublish first.",
        )
    q.deleted_at = datetime.now(tz=UTC)
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.put("/exams/{exam_id}/questions/order", response_model=list[QuestionOut])
async def reorder_questions(
    exam_id: uuid.UUID, body: ReorderIn, ctx: HrCtxDep, db: DbSessionDep
) -> list[QuestionOut]:
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    await _require_no_attempts(db, company_id, exam_id)

    live = await _live_questions(db, company_id, exam_id)
    if {q.id for q in live} != set(body.question_ids):
        raise HTTPException(
            status_code=400, detail="question_ids must be exactly the exam's live questions."
        )
    by_id = {q.id: q for q in live}
    # Two-phase to dodge the live (exam_id, position) partial-unique index: park
    # all rows at a high offset, flush, then assign the final 0..n-1 positions.
    for idx, qid in enumerate(body.question_ids):
        by_id[qid].position = 10_000 + idx
    await db.flush()
    now = datetime.now(tz=UTC)
    for idx, qid in enumerate(body.question_ids):
        by_id[qid].position = idx
        by_id[qid].updated_at = now
    await db.commit()
    return [_question_out(by_id[qid]) for qid in body.question_ids]


# ---------------------------------------------------------------------------
# Bulk add / AI generate / Excel import (locked once attempts exist)
# ---------------------------------------------------------------------------
@router.post(
    "/exams/{exam_id}/questions/bulk",
    status_code=status.HTTP_201_CREATED,
    response_model=list[QuestionOut],
)
async def bulk_add_questions(
    exam_id: uuid.UUID, body: BulkQuestionsIn, ctx: HrCtxDep, db: DbSessionDep
) -> list[QuestionOut]:
    """Append many questions at once (AI-generate 'add all', or a reviewed import)."""
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    await _require_no_attempts(db, company_id, exam_id)
    section = await _default_section(db, company_id, exam_id, "mcq")
    created = await _bulk_insert_questions(db, company_id, section, body.questions)
    log.info(
        "hr.exam.questions.bulk_added",
        exam_id=str(exam_id), company_id=str(company_id), count=len(created),
    )
    return [_question_out(q) for q in created]


@router.post("/exams/{exam_id}/questions/generate", response_model=GenerateQuestionsOut)
async def generate_questions(
    exam_id: uuid.UUID, body: GenerateQuestionsIn, ctx: HrCtxDep, db: DbSessionDep
) -> GenerateQuestionsOut:
    """Draft MCQs with Gemini (via feedback_billing). Returned for PREVIEW — not saved.
    HR reviews them, then persists the wanted ones via the bulk endpoint."""
    hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)

    try:
        raw = await generate_exam_questions_remote(
            topic=body.topic,
            num_questions=body.num_questions,
            difficulty=body.difficulty,
            language=body.language,
            acting_user_id=str(hr_uid),
        )
    except ExamGenerationError as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY, detail=f"AI generation failed: {exc}"
        ) from exc

    out: list[GeneratedQuestionOut] = []
    for q in raw:
        opts = [str(o) for o in (q.get("options") or [])]
        try:
            ci = int(q.get("correct_index", 0))
        except (TypeError, ValueError):
            continue
        prompt = str(q.get("prompt") or "").strip()
        if prompt and opts and 0 <= ci < len(opts):
            out.append(
                GeneratedQuestionOut(prompt=prompt, options=opts, correct_index=ci, points=1)
            )
    if not out:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="The AI returned no usable questions. Try again or refine the topic.",
        )
    log.info(
        "hr.exam.questions.generated",
        exam_id=str(exam_id), company_id=str(company_id), count=len(out),
    )
    return GenerateQuestionsOut(questions=out)


@router.post(
    "/exams/{exam_id}/questions/import",
    status_code=status.HTTP_201_CREATED,
    response_model=ImportQuestionsOut,
)
async def import_questions(
    exam_id: uuid.UUID,
    ctx: HrCtxDep,
    db: DbSessionDep,
    file: Annotated[UploadFile, File(description=".xlsx or .csv in the template layout")],
) -> ImportQuestionsOut:
    """Bulk-import questions from an uploaded Excel/CSV in the template layout.

    Valid rows are inserted; malformed rows are reported (partial success).
    """
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    await _require_no_attempts(db, company_id, exam_id)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    if len(content) > _MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="File too large (max 2 MB).")

    rows = _read_spreadsheet(file.filename or "", content)
    items, errors = _parse_question_rows(rows)
    if not items and not errors:
        raise HTTPException(
            status_code=400, detail="No question rows found. Use the template layout."
        )
    section = await _default_section(db, company_id, exam_id, "mcq")
    created = (
        await _bulk_insert_questions(db, company_id, section, items) if items else []
    )
    log.info(
        "hr.exam.questions.imported",
        exam_id=str(exam_id), company_id=str(company_id),
        added=len(created), errors=len(errors),
    )
    return ImportQuestionsOut(
        added=len(created),
        errors=errors,
        questions=[_question_out(q) for q in created],
    )


@router.get("/exam-question-template")
async def download_question_template(ctx: HrCtxDep) -> Response:
    """Download the .xlsx bulk-upload template (header + two example rows).

    Distinct top-level path (NOT /exams/...) so it never collides with the
    /exams/{exam_id} UUID route.
    """
    try:
        import openpyxl  # lazy import — only this route needs it
    except ImportError as exc:  # pragma: no cover - dep is in requirements
        raise HTTPException(
            status_code=500, detail="Excel support is not installed on the server."
        ) from exc
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(_TEMPLATE_HEADER)
    ws.append(["What is 2 + 2?", "3", "4", "5", "6", "B", "1"])
    ws.append(["Capital of France?", "Paris", "Rome", "Berlin", "Madrid", "A", "1"])
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="exam-questions-template.xlsx"'
        },
    )


# ---------------------------------------------------------------------------
# Assignments (magic-link mint / list / revoke)
# ---------------------------------------------------------------------------
@router.post(
    "/exams/{exam_id}/assignments", status_code=status.HTTP_201_CREATED, response_model=list[AssignOut]
)
async def assign_exam(
    exam_id: uuid.UUID, body: AssignIn, ctx: HrCtxDep, db: DbSessionDep
) -> list[AssignOut]:
    hr_uid, company_id = ctx
    exam = await _get_owned_exam(db, company_id, exam_id)

    # Resolve the target round (specified or the exam's first round). The token
    # grants exactly ONE round, and the ROUND is the unit that gets published +
    # scheduled separately — so the publish gate is on the round, not the exam.
    if body.round_id is not None:
        rnd = await db.scalar(
            select(ExamRound).where(
                ExamRound.id == body.round_id,
                ExamRound.exam_id == exam_id,
                ExamRound.company_id == company_id,
                ExamRound.deleted_at.is_(None),
            )
        )
        if rnd is None:
            raise HTTPException(status_code=404, detail="Round not found for this exam.")
    else:
        rnd = await _default_round(db, company_id, exam_id)

    if rnd.status != "published":
        raise HTTPException(status_code=409, detail="Publish this round before assigning it.")

    now = datetime.now(tz=UTC)
    if body.scheduled_at is not None and body.scheduled_at < now:
        raise HTTPException(status_code=422, detail="scheduled_at cannot be in the past.")

    ttl_hours = body.ttl_hours or settings.exam_link_ttl_hours
    expires_at = now + timedelta(hours=ttl_hours)
    base = settings.exam_link_base_url.rstrip("/")

    out: list[AssignOut] = []
    for applicant_id in body.applicant_ids:
        applicant = await db.scalar(
            select(Applicant).where(
                Applicant.id == applicant_id,
                Applicant.company_id == company_id,  # tenant scope: skip foreign applicants
                Applicant.deleted_at.is_(None),
            )
        )
        if applicant is None:
            continue  # not this company's applicant — silently skip

        # Rotate: revoke any existing active (invited) assignment for this
        # (round, applicant) so only one live link exists per round (the active
        # partial-unique index is now scoped to round_id).
        prior = await db.scalar(
            select(ExamAssignment).where(
                ExamAssignment.round_id == rnd.id,
                ExamAssignment.applicant_id == applicant_id,
                ExamAssignment.company_id == company_id,
                ExamAssignment.status == "invited",
                ExamAssignment.deleted_at.is_(None),
            )
        )
        if prior is not None:
            prior.status = "revoked"
            prior.updated_at = now
            await db.flush()

        raw_token = mint_exam_token()
        asn = ExamAssignment(
            id=uuid.uuid4(),
            company_id=company_id,
            exam_id=exam_id,
            round_id=rnd.id,
            applicant_id=applicant_id,
            created_by_user_id=hr_uid,
            token_hash=hash_exam_token(raw_token, settings.exam_link_secret),
            expires_at=expires_at,
            scheduled_at=body.scheduled_at,
            status="invited",
            created_at=now,
            updated_at=now,
        )
        db.add(asn)
        await db.flush()
        magic_link = f"{base}/exam#{raw_token}"  # raw token returned ONCE
        # Email the candidate their exam link (staged on this transaction →
        # atomic with the assignment, then delivered by the outbox worker). HR
        # still gets the link in the response to share manually if needed.
        await enqueue_email(
            db,
            to=applicant.email,
            template="exam_link",
            lang="en",
            ctx={
                "name": applicant.full_name,
                "exam_title": exam.title,
                "exam_url": magic_link,
                "when": (
                    body.scheduled_at.strftime("%d %b %Y, %H:%M UTC")
                    if body.scheduled_at else None
                ),
                "expires": expires_at.strftime("%d %b %Y, %H:%M UTC"),
            },
            company_id=company_id,
            related_kind="exam_assignment",
            related_id=asn.id,
        )
        out.append(
            AssignOut(
                assignment_id=str(asn.id),
                applicant_id=str(applicant_id),
                applicant_name=applicant.full_name,
                magic_link=magic_link,
                expires_at=expires_at.isoformat(),
                status=asn.status,
                round_id=str(rnd.id),
                scheduled_at=body.scheduled_at.isoformat() if body.scheduled_at else None,
            )
        )

    if not out:
        raise HTTPException(status_code=400, detail="No valid applicants in your company.")
    await db.commit()
    log.info(
        "hr.exam.assigned", exam_id=str(exam_id), company_id=str(company_id), count=len(out)
    )
    return out


@router.get("/exams/{exam_id}/assignments", response_model=list[AssignmentOut])
async def list_assignments(
    exam_id: uuid.UUID, ctx: HrCtxDep, db: DbSessionDep
) -> list[AssignmentOut]:
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    rows = (
        await db.execute(
            select(ExamAssignment, Applicant.full_name)
            .join(Applicant, Applicant.id == ExamAssignment.applicant_id)
            .where(
                ExamAssignment.exam_id == exam_id,
                ExamAssignment.company_id == company_id,
                ExamAssignment.deleted_at.is_(None),
            )
            .order_by(ExamAssignment.created_at.desc())
        )
    ).all()
    # An assignment past its expiry that was never completed is effectively dead
    # (the take endpoint already 404s it) — surface it as 'expired' rather than a
    # stale 'invited'/'started' so HR sees real link state.
    now = datetime.now(tz=UTC)
    return [
        AssignmentOut(
            assignment_id=str(a.id),
            applicant_id=str(a.applicant_id),
            applicant_name=name,
            status=(
                "expired"
                if a.status in ("invited", "started") and a.expires_at <= now
                else a.status
            ),
            expires_at=a.expires_at.isoformat(),
            consumed_at=a.consumed_at.isoformat() if a.consumed_at else None,
            created_at=a.created_at.isoformat(),
        )
        for a, name in rows
    ]


@router.post("/exams/{exam_id}/assignments/{aid}/revoke", response_model=AssignmentOut)
async def revoke_assignment(
    exam_id: uuid.UUID, aid: uuid.UUID, ctx: HrCtxDep, db: DbSessionDep
) -> AssignmentOut:
    _hr_uid, company_id = ctx
    asn = await db.scalar(
        select(ExamAssignment).where(
            ExamAssignment.id == aid,
            ExamAssignment.exam_id == exam_id,
            ExamAssignment.company_id == company_id,
            ExamAssignment.deleted_at.is_(None),
        )
    )
    if asn is None:
        raise HTTPException(status_code=404, detail="Assignment not found.")
    asn.status = "revoked"  # the token dies immediately on the take path
    asn.updated_at = datetime.now(tz=UTC)
    await db.commit()
    applicant_name = await db.scalar(
        select(Applicant.full_name).where(Applicant.id == asn.applicant_id)
    )
    return AssignmentOut(
        assignment_id=str(asn.id),
        applicant_id=str(asn.applicant_id),
        applicant_name=applicant_name or "",
        status=asn.status,
        expires_at=asn.expires_at.isoformat(),
        consumed_at=asn.consumed_at.isoformat() if asn.consumed_at else None,
        created_at=asn.created_at.isoformat(),
    )


# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------
@router.get("/exams/{exam_id}/attempts", response_model=list[AttemptResultOut])
async def list_attempts(
    exam_id: uuid.UUID,
    ctx: HrCtxDep,
    db: DbSessionDep,
    passed: Annotated[bool | None, Query()] = None,
    status_filter: Annotated[str | None, Query(alias="status")] = None,
) -> list[AttemptResultOut]:
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    stmt = (
        select(ExamAttempt, Applicant.full_name)
        .join(Applicant, Applicant.id == ExamAttempt.applicant_id)
        .where(
            ExamAttempt.exam_id == exam_id,
            ExamAttempt.company_id == company_id,
            ExamAttempt.deleted_at.is_(None),
        )
    )
    if passed is not None:
        stmt = stmt.where(ExamAttempt.passed.is_(passed))
    if status_filter:
        stmt = stmt.where(ExamAttempt.status == status_filter)
    stmt = stmt.order_by(ExamAttempt.score_percent.desc().nullslast(), ExamAttempt.submitted_at.desc())
    rows = (await db.execute(stmt)).all()
    return [
        AttemptResultOut(
            attempt_id=str(at.id),
            applicant_id=str(at.applicant_id),
            applicant_name=name,
            score_raw=at.score_raw,
            score_max=at.score_max,
            score_percent=at.score_percent,
            passed=at.passed,
            status=at.status,
            submitted_at=at.submitted_at.isoformat() if at.submitted_at else None,
            attempt_no=at.attempt_no,
        )
        for at, name in rows
    ]


@router.get("/exams/{exam_id}/attempts/{aid}/breakdown")
async def attempt_breakdown(
    exam_id: uuid.UUID, aid: uuid.UUID, ctx: HrCtxDep, db: DbSessionDep
) -> dict[str, Any]:
    """HR-ONLY per-question correctness, from the frozen graded_snapshot.

    This is the ONLY endpoint that exposes which answers were right — never the
    applicant take/submit path.
    """
    _hr_uid, company_id = ctx
    await _get_owned_exam(db, company_id, exam_id)
    at = await db.scalar(
        select(ExamAttempt).where(
            ExamAttempt.id == aid,
            ExamAttempt.exam_id == exam_id,
            ExamAttempt.company_id == company_id,
            ExamAttempt.deleted_at.is_(None),
        )
    )
    if at is None:
        raise HTTPException(status_code=404, detail="Attempt not found.")
    snapshot: dict[str, Any] = at.graded_snapshot or {}
    raw_answers: dict[str, Any] = at.answers or {}
    # New (round) attempts nest under 'mcq'/'coding'; legacy flat attempts are
    # {question_id: meta} with answers {question_id: index}. Detect + normalize.
    is_nested = bool(snapshot) and set(snapshot.keys()) <= {"mcq", "coding"}
    if is_nested:
        mcq_snapshot: dict[str, Any] = snapshot.get("mcq", {}) or {}
        mcq_answers_src = raw_answers.get("mcq", {}) if isinstance(raw_answers, dict) else {}
        coding_snapshot: dict[str, Any] = snapshot.get("coding", {}) or {}
    else:
        mcq_snapshot = snapshot
        mcq_answers_src = raw_answers
        coding_snapshot = {}
    answers: dict[str, int] = {k: int(v) for k, v in dict(mcq_answers_src).items()}
    questions = [
        GradeQuestion(
            question_id=qid,
            correct_index=int(meta.get("correct_index", -1)),
            points=int(meta.get("points", 1)),
        )
        for qid, meta in mcq_snapshot.items()
    ]
    per_question = grade_breakdown(GradeInput(questions=questions, answers=answers))
    return {
        "attempt_id": str(at.id),
        "score_percent": at.score_percent,
        "passed": at.passed,
        "per_question": per_question,
        "coding": coding_snapshot,
    }
